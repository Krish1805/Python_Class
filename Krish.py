# """ print("Hello")
# a=10
# b=30
# print(a == b)
# print(a != b)
# print( a>0 and a<50 )

# #--------------------

# user_name =input("Enter the Usename")
# Password =input("Enter the Password")

# Saved_user ="Krish"
# Saved_pass ="134"

# print( user_name == Saved_user and Password == Saved_pass ) """

# """ a=10
# a +=10
# print(a)
# """  """
# a= [1,"python",1.3]
# print(a)
# b= (1,"python",1.3)
# print (b)
# c ={'user':'Krish','pass':'13','mobile':'144444'}
# print(c.get('user'))
# print(c)
# d={'krish',1,'Hello'}
# print(d)
# print("python" in a) """

# #print(bin(9))
# #c ={'user':'Krish','pass':'13','mobile':'144444'}
# #print(c.get('user'))
# """ a=int(input("Enter the value:"))
# if a>10:
#     #print(a, "is greater than 10")
#     print(str(a) + " is greater than 10")
# elif a<10:
#     print(str(a) + " is less than 10") """
# """
# a=50

# while a>40:
#     print(a, "a is greater than 50")
#     a =a-1 """

# # a="python"
# # count=1
# # #start = int(input("Start Value:"))
# # #Last = int(input("Last Value:"))
# # #for j in range(start,Last +1)
# # #for j in range(start,Last +1,2):
# # for j in range(1,20):
# #     #count=1
# #     #print (j)
# #     #if(j%3==0):
# #         if count == 3:
# #             break
# #         print (j)
# #         count+=1

# # multiple commenting
# # a = """ Hello
# # world
# # welcome"""

# # print(a)

# # # Strings Indexing
# # a ="Python is a programming language"
# # # print(a[3])
# # # print(a[12:23])
# # # print(a[12:23:2]) # jumps 2 characters.This concept called as slicing
# # # print (a[-10:-21:-1]) # Negative indexing

# # # Strings are immuntable. Can not be changed once declared.
# # #a[0] ='p' # Error : 'str' object does not support item assignment.

# # #print(dir(str)) # prints medthods on string.
# # print ("Python" + "Django")
# # print("pthon"*3)

# # Sep 2nd : String methods

# # a="Krish"
# # b="python is a programming language"
# # print (a.startswith("k"))
# # print (a.endswith("h"))
# # print(a.isdigit())
# # print(b.count('p'))
# # print(b.index('i'))
# # print(b.index('p',1))
# # print (b.find('z'))

# # # Sep 3rd : String methods

# # a="python is a programming language"
# # #print(a.split())
# # #print(a[10:30].rsplit('a'))
# # b=["python","is","a","programming","language"]
# # #print(" ".join(b))
# # c=""
# # for j in b:
# #     c=" " + j
# #     print(c,end=" ")


# # Sep 4th ///// .format


# # name =input("Enter the name: " )
# # order_id =input("Enter the OrderID: ")
# # date =input("Enter the date: ")
# # #Option 1 # msg ="Hi {}, your order with OrderID {} has been successfully and it will be delivered by {}".format(name,OrderID,date)
# # # print(msg)
# # #Oprion 2(Keyname concept) # msg ="Hi,{name}, your order with OrderID{order_id} has been successfully and it will be delivred by {date}".format(order_id=order_id,date=date,name=name)
# # print(msg)

# # Value =int(input("Enter the Value: "))
# # print(Value)
# # j=2
# # while j<=Value:
# #     for i in range(1,j+1):

# #        numb=str(j)

# #        #num=num+num
# #        result="*".join(numb*j)
# #        print(result + "= " + str(j**j))
# #        j=j+1

# # rows = 6

# # for num in range(2,rows+1):

# #     for i in range(num):
# #         res ="*".join(str(num))
# #         res=res+res
# #         print(res)
# #         #print(str(num) ,end ="*") # print number
# #         #print(result)
# #     # line after each row to display pattern correctly

# #     print(" ")

# # Sep 7th

# a =[12,56.7,'python',[1,2,3,4],2]
# #print(a)
# #C:\Users\Nivi\Python
# #print(a[3][2])
# #print(a[1:4])
# #print(a[-1:-5:-1])
# #print(a[2][1:4])
# #print(a)
# #a[1]='Testing' # Lists are mutable.Values can be updated.
# #print(a)
# #a='Hello'
# #print(a)
# #a[5]='h'
# #del a[0]
# #print(a)
# #basic operations
# # a=[1,2,3,4]
# # b=[5,6,7,8,'python']
# # print(a+b)
# # print(b[4][2]*3)
# #### List methods
# #print(dir(list))
# # #help(list)
#  #a=[5,6,7,8,'python',[1,2,3,4],2]
# # #a.append(45)
# # #print(a)
# # b=[1,2,3]
# # a.append(b)
# # print(a)
# # a.extend(b)
# # print(a)
# # a.extend('45')
# # print(a)
# # a.insert(3,'Engineering')
# # print(a)
# ###### Sep 8th
# # a=[5,6,7,8,'python','python',[1,2,3,4],2]
# # a[6].extend(['Hello']) # entend can add elements to only inside lists.
# # #a[5].extend(44)
# # a.remove(a[0])
# # print(a)
# # #a.pop(4)
# # print(a)
# # #a.clear()
# # print(a)
# # print(a.index('python'))
# # print(a.count('python'))
# # b=[1,4,6,8,0]
# # b.sort()
# # print(b)
# # b.sort(reverse=True)
# # print(b)
# # b.reverse()
# # print(b)

# ########## Mutable and immutable concept
# # a=10
# # b=a
# # print(a)
# # print(b)
# # print(id(a) )
# # print(id(b))
# # print(type(a))
# # print(type(b))
# # a=10+1
# # print(id(a) )
# # print(id(b))
# # print(type(a))
# # print(type(b))
# # print(a)
# # print(b)
# # a=[1,2,3]
# # b=a
# # print(a)
# # print(b)
# # print(id(a) )
# # print(id(b))
# # print(type(a))
# # print(type(b))
# # a.pop()
# # print(a)
# # print(b)
# # print(id(a) )
# # print(id(b))
# # print(type(a))
# # print(type(b))

# ##### copy ()

# # a=[1,2,3,4]
# # b=a # creats copy in the same location
# # c=a.copy()# creats copy on diff location
# # a.append(5)
# # print(a)
# # print(b)
# # print(c)

# #### list comprehension

# # 1st way [expression for element in sequence ]

# # a= [10,20,30]
# # index = 0
# # for j in a:
# #     a[index]= j + 5
# #     index = index+1
# # print(a)

# # a= [10,20,30]
# # a=[j+5 for j in a]
# # print(a)

# #2nd way [expression for element in sequence  if condition]

# # a=[10,11,12,13,14,15]
# # b=[]
# # for j in a:
# #     if j%2 == 0:
# #         b.append(j)
# # print(b)


# # a=[10,11,12,13,14,15]
# # b=[j for j in a if j%2==0]
# # print(b)
# # print([j for j in a if j%2==0])


# ################# Sep 9th

# ## Tuple

# # a =(1,57.7,'python',[1,2,3],10,20,30)
# # print(type(a))
# # a=[12]
# # print(type(a))
# # a=(12)
# # print(type(a))
# # a=12, # single value tuple
# # print(type(a))
# # a=(12,)# single value tuple
# # print(type(a))
# # a=10,12,14 # to decclare tuple paranthesis () is not mandatory but recommended

# # a =(1,57.7,'python',[1,2,3],10,20,30)
# # #print(type(a))
# # print(a[0:len(a):2])

# # a=1,2,3,4
# # b=5,7,8
# # print(a+b)
# # print(a*3)
# # print(dir(tuple))

# #### Dictionaries ####
# # a={'name':'krish','id':'123'}
# # print(a)
# # print(a['name'])
# # a={"name":"krish","id":"123"}
# # print(a)
# # print(a["name"])
# # print(dir(dict))


# ###### Sep 10th

# # a={'name':'krish','id':'123'}
# # print(a['name'])
# # print(a.get('name'))
# # #print(a['users']) # Error as Users key is not found
# # print(a.get('users')) # Get() will return None when key value not found in dictionary instead of error.
# # print(a.get('users','Marketing')) # default value can be provided
# # print(a)

# # Nested dictionaries  accessing

# # a ={'emails':{'Primary':'Krish1@gmail.com','Secondary':'Krish2@gmail.com' },
# #     'passwords':{ 'Primarypass':'Pass1','Secondaypass':'pass2'  }
# #    }
# # print(a)
# # print(a['emails'])
# # print(a['passwords'])
# # print(a['emails']['Primary'])
# # print(a['passwords']['Primarypass'])


# # dictionaries are mutable
# # a={'name':'Krish','id':'12234'}
# # print(a)
# # a['name'] ='Krishna' #updates Key value
# # a['mobile'] ='878888' # add new key:value
# # print(a)

# # dictionary methods

# # a={'name':'Krish','id':'12234'}
# # b={'addr':'Hyderabad','State':'Telanagana'}

# ## manual logic

# # for j in b:
# #     a[j]=b[j]
# # print(a)

# #a.update(b) # adds elements from b to a
# #print(a)

# #a.pop('id') #removes id key value
# #a.popitem() # removes last key :value by default
# #a.clear()

# # print(a)
# # print(a.keys())
# # print(a.values())
# # print(a.items())


# # 11th Sep

# # a=['name','id','mobile']
# # b={}

# # for j in a: #mamnual way to create dynamic dictionary
# #     b[j]=None
# # print(b)

# #print({}.fromkeys(a))

# # a ={'name':'krish'}
# # b=a
# # #a.setdefault('password')
# # #print(a)
# # a.setdefault('password','123')
# # print(a)
# # a.update({'password':'4455'})
# # print(a)

# #a ={'name':'krish'}
# #b=a # uses same memmory
# #a.update({'name':'krishna'})
# # b=a.copy()
# # print(a)
# # print(b)
# # print(id(a))
# # print(id(b))

# # a.update({'name':'krishna'})
# # print(a)
# # print(b)
# # print(id(a))
# # print(id(b))

# #### Dictionary comprehension

# # ## 1st syntax 

# # #{expression for element in sequence}

# # a={}
# # # for j in range(1,5):
# # #     a[j] =j*2
# # # print(a)
# # print({j:j*2 for j in range(1,5)})

# ## 1st syntax 
# #{expression for element in sequence}


# ### 3rd syntax -Condition
# #{expression  for element in sequence if condition}

# #a={}
# # for j in range(1,10):
# #     if j%2 ==0:
# #         a[j] =j*2
# # print(a)
# #print({j:j*2 for j in range(1,10) if j%2==0})


# ### 3rd syntax -Condition
# #{expression  if condition else value for element in sequence }

# # a={}
# # for j in range(1,5):
# #     if j%2 ==0:
# #         a[j] =j*2
# #     else:
# #         a[j] =j*3
# # print(a)
# #print({j:j*2 if j%2==0 else j*3 for j in range(1,10)})

# # Sep 15th 

# #Sets
# # sets will have unique values
# # order of elements is constant

# # a={1,2,3,4,4,'python'}
# # print(a)
# # a={}
# # print(type(a))
# # a=set([]) # to define empty set
# # print(type(a))

# #print(dir(set))

# # a={19,20,'python',8.5}
# # a.add('Hello')
# # print(a)
# # #a.add([1,'testing']) # Error. mutable elements can not be added.
# # a.update([1,'testing']) # it will add elements as individual elements
# # a.remove('Hello')
# # print(a)
# # a.discard('hello') # does not throw error when element to be removed is not found in set
# # print(a)
# # a.pop() # random element is removed.
# # print(a)
# # a.clear()
# # print(a)

# # Union -  combines multiple sets and displace unique elements
# # a={10,20,30}
# # b={10,40,50}
# # print(a|b)
# # print(a.union(b))
# # # intersection(&)  # common elements

# # print(a.intersection(b))
# # print(a&b)

# # #difference - returns common element in first set and returns rest elements from first set
# # print(a-b)
# # print(a.difference(b))
# # # symmertic_difference
# # print(a^b) # retunrs uncommon elements from both sets
# # print(a.symmetric_difference(b))


# # a={10,20,30}
# # b={10,40,50}
# # print(a)
# # #a.intersection_update(b) # changes a set with common values
# # #a.difference_update(b)
# # a.symmetric_difference_update(b)
# # print(a)



# # a={10,20,30}
# # b={10,40,50}
# # print(a.isdisjoint(b)) # checks if sets are idsjoint sets( no common elements from both sets)

# ##### Sep 17th  - Module 3

# # Functions - Set of lines of code used to perform specific task. mostly used for code reusability

# # a=75
# # b=55
# # def arthemetic():
# #     print(a+b)
# #     print(a-b)
# #     print(a*b)
# #     print(a/b)

# # arthemetic()

# # Passing arguments to fucntion can be done in 5 ways

# ## Postional arguments

# # def arthemetic(a,b):
# #     print(a+b)
# #     print(a-b)
# #     print(a*b)
# #     print(a/b)
# # arthemetic(25,50)

# # def std_data(name,email):
# #     print("{} is having {} email id ".format(name,email))
# # std_data('krish','krishna@gmail.com')

# ## default arguments

# # def std_data(name,email="123@gmail.com"):
# #     print("{} is having {} email id ".format(name,email))
# # std_data('krish','krishna@gmail.com')
# # std_data('krish')

# ## Sep 18th 

# # def std_data(name ='krishna',email): # Error. Non defaults arguments can not be followed by defualt arguments
# #     print("{} is having {} email id ".format(name,email))
# # std_data('krish','krishna@gmail.com')
# # #std_data('krish')

# ### Keyboard arguments --Providing the value based on the key name at fucntion call

# # def std_data(name ,email): 
# #     print("{} is having {} email id ".format(name,email))
# # std_data(name ='krish',email ='krishna@gmail.com')
# # std_data(email ='123@gmail.com',name ='krish')
# # std_data('krish',email ='krishna@gmail.com')
# # #std_data(email ='123@gmail.com','krish') # Error

# # Arbitary arguments

# # def addition(*a):
# #     print(a)
# #     print(type(a))
# # addition(1,2)
# # addition(1,2,3)

# # def bill_generation(*bill):
# #     print(bill)
# #     final_bill = 0
# #     for j in bill[:-1]:
# #         final_bill = final_bill+j
# #     # print(final_bill)
# #     print("{} amount is spent in this month from card with {} number..".format(final_bill,bill[-1]))

# # bill_generation(1253,2345,7483,1256,'HDFC1234')

# #### Sep 22nd

# #Arbitary keyboard arguments

# # def addition(**value):
# #     print(value)
# # addition(a=10,b=20)
# # addition(x=10,y=20,z=35)

# # def addition(**value):
# #     print(value)
# #     count=0
# #     for j in value:
# #         count= count + value[j]
# #     print(count)
# # addition(x=10,y=20,z=35)


# # def year_tran(accno,**trans):
# #     print(trans)
# #     final_bill=0
# #     for j in trans:
# #         final_bill =final_bill + trans[j]
# #     print(" Your account ID with {} has been charged {} for this year.".format(accno,final_bill) )
# # year_tran("HDFC1501234",jan=123,feb=7878)


# #### Return statement
# # def add(a,b):
# #     #print(a+b)
# #     return a+b
# # #print(add(4,5))

# # def sub(b):
# #     print(add(4,5) -b)
# # sub(6)

# ## Recursion

# # def cal_fac(n):
# #     if n==1:
# #         return 1
# #     else:
# #         return n *cal_fac(n-1)
# # n=int(input("enter the value:"))
# # print(cal_fac(n))

# ### 23rd Sep

# #Global and local Variables


# # a=10
# # def add(b,c):
# #     print(a+b+c)
# # add(20,30)
# # print (b) # Error.This is local variable

# # # Lambda fucntion

# # x =lambda a:a*2
# # print(x(5))

# # # alternative way
# # print ((lambda a:a*2)(5))


# # def sample(a):
# #     return a*2
# # print(sample(5))

# # a=[]
# # def sample(b):
# #     for j in b:
# #         a.append(j*2)
# # sample([10,13,24])
# # print(a)

# # alternate way using map

# # print(map(lambda a:a*2,[10,13,24]))
# # print(list(map(lambda a:a*2,[10,13,24])))

# ## filter for lambda
# #print(list(filter( lambda a:a%2==0,[5,23,56,13,14])))

# # a=['900001','7000001','911101']
# # #c=list(filter(lambda a:a.startswith('9'),a))
# # #print(list(filter(lambda a:a.endswith('1'),c)))
# # # alternate way
# # print(list(filter(lambda a:a.endswith('1'),list(filter(lambda a:a.startswith('9'),a)))))



# ### 24th Sep  - Modules -- Any python file with .py extension

# # Userdefined modules --User created
# #inbuilt modules -- Python modules
# #external modules -- used from external

# #help('modules') -- to list down modules


# # import module1 # imports complete module
# # print(module1.a)
# # print(module1.add(25,35))

# # from module1 import * # imports funtionalities of module
# # print(a)
# # print(add(45,35))

# ## Inoder to access other modules from current modules(modules located at different locations), python will look for below paths
# #1) current working folder
# #2) folder where python is installed
# #3) Python environmental variables

# # to know where python can import modules

# # import sys
# # print(sys.path)
# # sys.path.append('C:/Users/Nivi/Documents/My Tableau Repository') # to add new path
# # print(sys.path)


# # 25th Sep
# # Inbuilt modules - Modules that comes with python
# #print(help('modules')) # to list down list of modules
# # import math
# # #print(help(math))
# # print(math.exp(4))
# # print(math.pow(4,3))

# # #random()
# # import random
# # print(random.random()) # random value between 0 to 1
# # print(random.randint(10000,99999))
# # a=["SWIGGY40",'SPECIAL30',"SWIGGYIT","ZOMATO"]
# # print(random.choice(a))
# # print(random.sample(a,2))
# # random.shuffle(a)
# # print(a)

# # ### os module
# # import os
# # print(os.getcwd())
# # os.chdir('C:/Users/Nivi/python')
# # print(os.getcwd())
# # print(os.listdir('.'))
# # print(os.listdir('C:/Users/Nivi/python'))

# # 28th Sep
# # import os
# # # print(os.listdir())
# # # print(os.listdir('C:/Users/Nivi/python'))# list only current folder list
# # # print(list(os.walk('C:/Users/Nivi/python'))) # includes subfolders list as well
# # #os.mkdir('C:/Users/Nivi/python/PythonTest')
# # print(list(os.walk('C:/Users/Nivi/python'))) 
# # #os.makedirs('folder1/folder2')
# # print(list(os.walk('C:/Users/Nivi/python'))) 
# # os.removedirs('folder1/folder2')
# # print(list(os.walk('C:/Users/Nivi/python'))) 

# #### datetime()  module
# #import datetime
# # print(datetime.datetime.now())
# # print(datetime.datetime.now().date())
# # print(datetime.datetime.now().time())
# # print(datetime.datetime.now().day)
# # today_date=datetime.datetime.now()
# # after_30days=datetime.timedelta(days=30)
# # # print(today_date+after_30days)
# # # help(datetime)

# # #datetime to string coversion
# # sample_dt =(today_date+after_30days).date()
# # print(sample_dt)
# # #strf
# # print(sample_dt.strftime('%Y-%m-%d'))
# # print(type(sample_dt.strftime('%Y-%m-%d'))) #String format

# # 29th Sep
# #strftime  - converting datetime into string
# #strptime- coverting string into datatime

# # a='2020-01-30'
# # x='2020-01-30 09:09:09'
# # b=datetime.datetime.strptime(a,'%Y-%m-%d')
# # c=datetime.datetime.strptime(x,'%Y-%m-%d %H:%M:%S')
# # print(b)
# # print(c)
# # print(type(b))
# # print(c.date())
# # print(c.time())

# # m=datetime.datetime(1988,6,18)
# # print(m)
# # print((m.strftime('%d-%m-%Y %A %B %b')))

# # Calender Module

# # import calendar
# # print(calendar.firstweekday())
# # print(list(calendar.day_name))
# # print(calendar.day_name[0])
# # print(calendar.day_abbr[0])
# # #help('modules')

# # print(calendar.isleap(2016))
# # print(calendar.month(2021,1))
# # print(calendar.calendar(2021))

# # # urllib

# # import urllib.request
# # b=urllib.request.urlopen('https://www.flipkart.com/nike-revolution-5-running-shoes-men/p/itm11637de31a718?pid=SHOFK455XMBK3PYE&lid=LSTSHOFK455XMBK3PYER3OTAW&marketplace=FLIPKART&srno=b_1_1&otracker=hp_omu_Deals%2Bof%2Bthe%2BDay_2_4.dealCard.OMU_ZDWY56YXQIG6_3&otracker1=hp_omu_SECTIONED_neo%2Fmerchandising_Deals%2Bof%2Bthe%2BDay_NA_dealCard_cc_2_NA_view-all_3&fm=neo%2Fmerchandising&iid=2640436c-6b96-467d-bd32-1039b4e6b212.SHOFK455XMBK3PYE.SEARCH&ppt=browse&ppn=browse&ssid=66zd3x7o740000001612027081512')
# # #print(b.read())
# # flipkart_data=b.read()
# # with open('flipkart.html','w') as file_data:
# #     file_data.write(str(flipkart_data))

# #Beautifulsoup(bs4)

# # Sep 30th 

# # Packages

# # from Testfolder import Sample
# # print(Sample.squart(25))
# # from PythonTest.Krishna import Test
# # print(Test.math.pow(4,3))

# # git - code version controller

# Oct 7th and # Oct 8th

# conencting to mysql

# import pymysql
# #print(pymysql.connect(user='root',password='',host='localhost'))
# conn =pymysql.connect(user='root',password='',host='localhost')
# cur=conn.cursor()
# #cur.execute('create database python')
# #cur.execute('create table python.Student(Id int(10),name varchar(255))')
# cur.execute('use python')
# cur.execute(" insert into student(id,name) values(2,'krishna')")
# cur.execute('commit')
# cur.execute('select * from student')
# print(cur.fetchall())
# #cur.execute('show tables')
# #print(cur.execute('show tables'))
# print(cur.fetchall())
# conn.close()

# fileoperations
# file_data = open('Sample.txt','r')
# print(file_data)
# file_data.close()
# print(file_data.closed)
# with open('Sample.txt','r') as file_data:
#     #for j in file_data.readline():
#     #for j in file_data.readlines():
#     #for j in file_data.read():
#         print(file_data.readlines())

# OCt 9th
# with open('Sample.txt','r') as file_data:
#     data=file_data.readlines()
#     for line in data:
#         if "Krish" in line: 
#             print(line)
# Writing 

# with open('Writefile.txt','w') as file_data:
#     #file_data.write('My Seconnd Write operation\n')
#     file_data.writelines(['Fist Line\n','Second line\n','Third line\n'])

# with open('Writefile.txt','a') as file_data:
#     file_data.write('My Seconnd Write operation\n')
#     file_data.writelines(['Fist Line\n','Second line\n','Third line\n'])


# csv module

# import csv
# with open('Student.csv','r') as file_data:
#     csv_reader=csv.reader(file_data)
#     next(csv_reader)# skips one row at a time
#     #print(csv_reader)
#     for j in csv_reader:
#         if j[2] == "ECE" and j[1].startswith('K'):
#             print(j)

### OCt 12th and 13th 
import csv
#with open('Student_details.csv','w') as file_data:
# with open('Student_details.csv','w',newline='') as file_data:# removes extra lines
#     csv_writer=csv.writer(file_data)
#     csv_writer.writerow(['name','Id','Dept'])
#     csv_writer.writerow(['Krish','10','ECE'])
#     csv_writer.writerows([['Lokesh','20','CSV'],['Nivi','30','IT']])
# with open('Student_details.csv','a',newline='') as file_data:# removes extra lines
#     csv_writer=csv.writer(file_data)
#     csv_writer.writerow(['name','Id','Dept'])
# #     csv_writer.writerow(['Krish','10','ECE'])
# #     csv_writer.writerows([['Lokesh','20','CSV'],['Nivi','30','IT']])
# with open('Student_details.csv','r') as file_data:
#     csv_reader=csv.reader(file_data)
#     read_data=[]
#     csv_reader =list(csv_reader)
#     #next(csv_reader)
# #with open('Student_details.csv','a',newline='') as file_data_W:
#     #csv_writer=csv.writer(file_data_W)
#     for i in range(0,len(csv_reader)) : 
#         if i%2!=0:
#             read_data.append(csv_reader[i])
# with open('Write2.csv','a',newline='') as file_data_w:
#     csv_writer=csv.writer(file_data_w)
#     csv_writer.writerows(read_data)

# Excel Operations
# 
#pip install openpyxl
# import openpyxl
# #print(dir(openpyxl))       
# data =openpyxl.load_workbook('Employee_details.xlsx') 
# # print(data)
# # print(data.sheetnames)
# #print(data.get_sheet_by_name('Student_details'))
# datasheet= data['Department']
# #print(datasheet)
# # print(datasheet['B3'].value) # 1 way to pick cell value
# # print(datasheet.cell(row=3,column=2).value) # 2nd way to pick cell value
# # print(datasheet.max_row)
# # print(datasheet.max_column)
# row_count=datasheet.max_row
# coulumn_count=datasheet.max_column
# for j in range(1,coulumn_count+1):
#     print(datasheet.cell(row=2,column=j).value,end=" ")
# for j in range(1,row_count+1):
# #     print(datasheet.cell(row=j,column=2).value,end=" ")
# for j in range(1,row_count+1):
#     for i in range(1,coulumn_count+1):
#         if datasheet.cell(row=j,column=2).value.startswith('H'):
#             print(datasheet.cell(row=j,column=i).value,end=" ")
#         else:
#             break
#     print()

# Writing data into Excel    

#import openpyxl
# work_book=openpyxl.Workbook()
# work_book.create_sheet('Employee')
# sheet1=work_book['Sheet']
# work_book.remove_sheet(sheet1)
# sheet_data=work_book['Employee']
# #sheet_data.append(['ID','Name','Dept'])
# data=[
#     ['Id','Name','Dept'],
#     ['Krish',10,'HR'],
#     ['Raj',20,'Finance']
# ]
# for j in data:
#     sheet_data.append(j)

# work_book.save('Write_data.xlsx')

#Oct 16th 

####### Appending data into excel #######

import openpyxl
data= openpyxl.load_workbook('Write_data.xlsx')
datasheet=data['Employee']
row_count=datasheet.max_row
coulumn_count=datasheet.max_column
out_list=[]
for j in range(2,row_count+1):
    in_list=[]
    for i in range(1,coulumn_count+1):
            in_list.append(datasheet.cell(row=j,column=i).value)
        
    out_list.append(in_list)

work_book=openpyxl.Workbook()
work_book.create_sheet('Employee')
sheet1=work_book['Sheet']
work_book.remove_sheet(sheet1)
sheet_data=work_book['Employee']
#sheet_data.append(['ID','Name','Dept'])
data=[
    ['Nivi2',10,'HR'],
    ['Lokesh2',20,'Finance']
]
out_list.extend(data)
print(out_list)
for j in out_list:
    sheet_data.append(j)
    
work_book.save('Write_data.xlsx')

####### Appending data into excel -END #######


















